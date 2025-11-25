"""
Utilidades para generación de reportes académicos con gráficos
Incluye funciones para crear gráficos en base64 para HTML/PDF y para Excel
"""
import io
import base64
from datetime import datetime
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')  # Backend sin interfaz gráfica para servidores
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import numpy as np

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.db import models
from django.db.models import Avg, Count, Q, Max, Min, F, Case, When, IntegerField, FloatField
from django.db.models.functions import ExtractMonth
from academico.models import (
    Alumno, Materia, Comision, InscripcionAlumnoComision,
    Calificacion, Asistencia, TipoCalificacion, AnioAcademico
)


def generar_grafico_base64(fig):
    """
    Convierte una figura de matplotlib a base64 para incrustar en HTML o PDF
    """
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
    plt.close(fig)
    return image_base64


def grafico_distribucion_notas(calificaciones_data, titulo="Distribución de Notas"):
    """
    Genera un gráfico de barras con la distribución de notas
    calificaciones_data: lista de tuplas (materia_nombre, nota_promedio)
    """
    if not calificaciones_data:
        return None

    fig, ax = plt.subplots(figsize=(10, 6))

    materias = [item[0][:20] for item in calificaciones_data]  # Limitar nombre
    notas = [item[1] for item in calificaciones_data]

    colores = ['#2ecc71' if nota >= 6 else '#e74c3c' for nota in notas]

    bars = ax.bar(materias, notas, color=colores, alpha=0.7)
    ax.axhline(y=6, color='#f39c12', linestyle='--', linewidth=2, label='Nota mínima (6)')

    ax.set_xlabel('Materias', fontsize=12)
    ax.set_ylabel('Nota Promedio', fontsize=12)
    ax.set_title(titulo, fontsize=14, fontweight='bold')
    ax.set_ylim(0, 10)
    ax.legend()
    ax.grid(axis='y', alpha=0.3)

    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    return generar_grafico_base64(fig)


def grafico_aprobados_desaprobados(aprobados, desaprobados, regulares):
    """
    Genera un gráfico circular con distribución de estados académicos
    """
    if aprobados == 0 and desaprobados == 0 and regulares == 0:
        return None

    fig, ax = plt.subplots(figsize=(8, 8))

    labels = []
    sizes = []
    colors = []

    if aprobados > 0:
        labels.append(f'Aprobados ({aprobados})')
        sizes.append(aprobados)
        colors.append('#2ecc71')

    if desaprobados > 0:
        labels.append(f'Desaprobados ({desaprobados})')
        sizes.append(desaprobados)
        colors.append('#e74c3c')

    if regulares > 0:
        labels.append(f'Regulares ({regulares})')
        sizes.append(regulares)
        colors.append('#f39c12')

    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels,
        colors=colors,
        autopct='%1.1f%%',
        startangle=90,
        textprops={'fontsize': 12}
    )

    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')

    ax.set_title('Distribución de Estados Académicos', fontsize=14, fontweight='bold')
    plt.tight_layout()

    return generar_grafico_base64(fig)


def grafico_evolucion_asistencias(asistencias_por_mes):
    """
    Genera un gráfico de líneas con la evolución de asistencias
    asistencias_por_mes: dict {mes: porcentaje_asistencia}
    """
    if not asistencias_por_mes:
        return None

    fig, ax = plt.subplots(figsize=(10, 6))

    meses = list(asistencias_por_mes.keys())
    porcentajes = list(asistencias_por_mes.values())

    ax.plot(meses, porcentajes, marker='o', linewidth=2, markersize=8, color='#3498db')
    ax.fill_between(meses, porcentajes, alpha=0.3, color='#3498db')
    ax.axhline(y=75, color='#f39c12', linestyle='--', linewidth=2, label='Asistencia mínima (75%)')

    ax.set_xlabel('Mes', fontsize=12)
    ax.set_ylabel('Porcentaje de Asistencia (%)', fontsize=12)
    ax.set_title('Evolución de Asistencias', fontsize=14, fontweight='bold')
    ax.set_ylim(0, 100)
    ax.legend()
    ax.grid(True, alpha=0.3)

    plt.xticks(rotation=45)
    plt.tight_layout()

    return generar_grafico_base64(fig)


def grafico_comparativo_alumnos(alumnos_data, metrica='promedio'):
    """
    Genera un gráfico comparativo de múltiples alumnos
    alumnos_data: lista de tuplas (apellido_nombre, valor)
    """
    if not alumnos_data:
        return None

    fig, ax = plt.subplots(figsize=(12, 6))

    nombres = [item[0][:25] for item in alumnos_data]
    valores = [item[1] for item in alumnos_data]

    colores = plt.cm.viridis(np.linspace(0, 1, len(nombres)))

    bars = ax.barh(nombres, valores, color=colores, alpha=0.7)

    ylabel_map = {
        'promedio': 'Nota Promedio',
        'asistencia': 'Porcentaje de Asistencia (%)',
        'materias_aprobadas': 'Cantidad de Materias Aprobadas'
    }

    ax.set_xlabel(ylabel_map.get(metrica, 'Valor'), fontsize=12)
    ax.set_ylabel('Alumnos', fontsize=12)
    ax.set_title(f'Comparativo de Alumnos - {ylabel_map.get(metrica, "Métrica")}',
                 fontsize=14, fontweight='bold')
    ax.grid(axis='x', alpha=0.3)

    plt.tight_layout()

    return generar_grafico_base64(fig)


def obtener_datos_reporte_academico(filtros=None):
    """
    Obtiene y procesa datos para el reporte académico completo (OPTIMIZADO)

    filtros: dict con opciones {
        'comision_id': int,
        'materia_id': int,
        'anio_academico': int,
        'fecha_inicio': date,
        'fecha_fin': date
    }

    Retorna: dict con todos los datos procesados y listos para gráficos
    """
    filtros = filtros or {}

    # VALIDACIÓN: Al menos un filtro debe estar presente para evitar sobrecarga
    if not any([filtros.get('comision_id'), filtros.get('materia_id'), filtros.get('anio_academico')]):
        # Sin filtros, usar solo el año académico activo
        anio_activo = AnioAcademico.objects.filter(activo=True).first()
        if anio_activo:
            filtros['anio_academico'] = anio_activo.id

    # Base queryset de inscripciones
    inscripciones = InscripcionAlumnoComision.objects.select_related(
        'alumno', 'comision', 'comision__materia', 'comision__anio_academico'
    )

    # Aplicar filtros
    if filtros.get('comision_id'):
        inscripciones = inscripciones.filter(comision_id=filtros['comision_id'])

    if filtros.get('materia_id'):
        inscripciones = inscripciones.filter(comision__materia_id=filtros['materia_id'])

    if filtros.get('anio_academico'):
        inscripciones = inscripciones.filter(comision__anio_academico=filtros['anio_academico'])

    # 1. DATOS DE ALUMNOS (con límite)
    total_alumnos = inscripciones.values('alumno').distinct().count()

    # 2. CALIFICACIONES POR MATERIA (OPTIMIZADO con agregación)
    # Limitar a top 15 materias para evitar gráficos ilegibles
    promedios_materias_query = Calificacion.objects.filter(
        alumno_comision__in=inscripciones
    ).values(
        'alumno_comision__comision__materia__nombre'
    ).annotate(
        promedio=Avg('nota')
    ).order_by('-promedio')[:15]

    promedios_materias = [
        (item['alumno_comision__comision__materia__nombre'], item['promedio'])
        for item in promedios_materias_query
    ]

    # 3. ESTADOS ACADÉMICOS (OPTIMIZADO con agregación)
    # Calcular estados basándose en:
    # 1. Si tiene nota FINAL: usar esa nota
    # 2. Si NO tiene nota FINAL pero tiene otras notas: calcular promedio
    # 3. Si NO tiene notas: considerarlo "Regular/En Curso"

    # Inscripciones con nota FINAL
    inscripciones_con_final = Calificacion.objects.filter(
        alumno_comision__in=inscripciones,
        tipo=TipoCalificacion.FINAL
    ).values('alumno_comision_id').annotate(
        nota_final=Avg('nota')
    )

    ids_con_final = [item['alumno_comision_id'] for item in inscripciones_con_final]
    aprobados_final = sum(1 for item in inscripciones_con_final if item['nota_final'] >= 6)
    desaprobados_final = sum(1 for item in inscripciones_con_final if item['nota_final'] < 6)

    # Inscripciones SIN nota final pero CON otras notas (usar promedio)
    inscripciones_sin_final = Calificacion.objects.filter(
        alumno_comision__in=inscripciones
    ).exclude(
        alumno_comision_id__in=ids_con_final
    ).values('alumno_comision_id').annotate(
        promedio=Avg('nota')
    )

    aprobados_promedio = sum(1 for item in inscripciones_sin_final if item['promedio'] >= 6)
    desaprobados_promedio = sum(1 for item in inscripciones_sin_final if item['promedio'] < 6)

    # Total de aprobados y desaprobados
    aprobados = aprobados_final + aprobados_promedio
    desaprobados = desaprobados_final + desaprobados_promedio

    # Regulares/En curso: inscripciones sin ninguna calificación
    ids_con_calificaciones = set(ids_con_final + [item['alumno_comision_id'] for item in inscripciones_sin_final])
    regulares = inscripciones.exclude(id__in=ids_con_calificaciones).count()

    # 4. ASISTENCIAS POR MES (OPTIMIZADO con agregación)
    asistencias_query = Asistencia.objects.filter(
        alumno_comision__in=inscripciones
    )

    if filtros.get('fecha_inicio'):
        asistencias_query = asistencias_query.filter(fecha_asistencia__gte=filtros['fecha_inicio'])
    if filtros.get('fecha_fin'):
        asistencias_query = asistencias_query.filter(fecha_asistencia__lte=filtros['fecha_fin'])

    # Usar SQL agregado en lugar de iterar
    from django.db.models.functions import TruncMonth, ExtractMonth
    from django.db.models import Case, When, IntegerField

    asistencias_mes = asistencias_query.annotate(
        mes=ExtractMonth('fecha_asistencia')
    ).values('mes').annotate(
        total=Count('id'),
        presentes=Count(Case(When(esta_presente=True, then=1), output_field=IntegerField()))
    ).order_by('mes')

    # Mapeo de números de mes a nombres
    meses_nombres = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
                     7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}

    porcentajes_por_mes = {
        meses_nombres[item['mes']]: (item['presentes'] / item['total'] * 100) if item['total'] > 0 else 0
        for item in asistencias_mes
    }

    # 5. TOP ALUMNOS (OPTIMIZADO con una sola query por métrica)
    # Top 10 alumnos por promedio
    alumnos_promedios_query = Calificacion.objects.filter(
        alumno_comision__in=inscripciones
    ).values(
        'alumno_comision__alumno__apellido',
        'alumno_comision__alumno__nombre'
    ).annotate(
        promedio=Avg('nota')
    ).order_by('-promedio')[:10]

    alumnos_promedios = [
        (f"{item['alumno_comision__alumno__apellido']} {item['alumno_comision__alumno__nombre']}",
         item['promedio'])
        for item in alumnos_promedios_query
    ]

    # Top 10 alumnos por asistencia
    alumnos_asistencias_query = Asistencia.objects.filter(
        alumno_comision__in=inscripciones
    ).values(
        'alumno_comision__alumno__apellido',
        'alumno_comision__alumno__nombre'
    ).annotate(
        total=Count('id'),
        presentes=Count(Case(When(esta_presente=True, then=1), output_field=IntegerField()))
    ).annotate(
        porcentaje=Case(
            When(total__gt=0, then=100.0 * models.F('presentes') / models.F('total')),
            default=0,
            output_field=models.FloatField()
        )
    ).order_by('-porcentaje')[:10]

    alumnos_asistencias = [
        (f"{item['alumno_comision__alumno__apellido']} {item['alumno_comision__alumno__nombre']}",
         item['porcentaje'])
        for item in alumnos_asistencias_query
    ]

    # Top 10 alumnos por materias aprobadas
    alumnos_materias_query = Calificacion.objects.filter(
        alumno_comision__in=inscripciones,
        tipo=TipoCalificacion.FINAL,
        nota__gte=6
    ).values(
        'alumno_comision__alumno__apellido',
        'alumno_comision__alumno__nombre'
    ).annotate(
        aprobadas=Count('id')
    ).order_by('-aprobadas')[:10]

    alumnos_materias_aprobadas = [
        (f"{item['alumno_comision__alumno__apellido']} {item['alumno_comision__alumno__nombre']}",
         item['aprobadas'])
        for item in alumnos_materias_query
    ]

    # 6. ESTADÍSTICAS GENERALES (OPTIMIZADO)
    total_materias = inscripciones.values('comision__materia').distinct().count()
    total_comisiones = inscripciones.values('comision').distinct().count()

    promedio_general = Calificacion.objects.filter(
        alumno_comision__in=inscripciones
    ).aggregate(Avg('nota'))['nota__avg'] or 0

    # Asistencia general
    asist_stats = asistencias_query.aggregate(
        total=Count('id'),
        presentes=Count(Case(When(esta_presente=True, then=1), output_field=IntegerField()))
    )
    porcentaje_asistencia_general = (asist_stats['presentes'] / asist_stats['total'] * 100) if asist_stats['total'] > 0 else 0

    return {
        'filtros_aplicados': filtros,
        'fecha_generacion': datetime.now(),

        # Estadísticas
        'estadisticas': {
            'total_alumnos': total_alumnos,
            'total_materias': total_materias,
            'total_comisiones': total_comisiones,
            'promedio_general': round(promedio_general, 2) if promedio_general else 0,
            'porcentaje_asistencia_general': round(porcentaje_asistencia_general, 2),
            'aprobados': aprobados,
            'desaprobados': desaprobados,
            'regulares': regulares,
        },

        # Datos para gráficos
        'promedios_materias': promedios_materias,
        'estados_academicos': (aprobados, desaprobados, regulares),
        'asistencias_por_mes': dict(porcentajes_por_mes),
        'alumnos_top_promedio': alumnos_promedios[:10],
        'alumnos_top_asistencia': alumnos_asistencias[:10],
        'alumnos_materias_aprobadas': alumnos_materias_aprobadas[:10],

    }


def generar_excel_reporte_academico(datos_reporte):
    """
    Genera un archivo Excel con datos y gráficos del reporte académico
    """
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    # HOJA 1: RESUMEN
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"

    # Título
    ws_resumen['A1'] = 'REPORTE ACADÉMICO COMPLETO'
    ws_resumen['A1'].font = Font(bold=True, size=16)
    ws_resumen['A2'] = f"Generado: {datos_reporte['fecha_generacion'].strftime('%d/%m/%Y %H:%M')}"

    # Estadísticas
    row = 4
    ws_resumen[f'A{row}'] = 'ESTADÍSTICAS GENERALES'
    ws_resumen[f'A{row}'].font = Font(bold=True, size=14)

    estadisticas = datos_reporte['estadisticas']
    row += 2
    for key, value in estadisticas.items():
        ws_resumen[f'A{row}'] = key.replace('_', ' ').title()
        ws_resumen[f'B{row}'] = value
        row += 1

    # HOJA 2: PROMEDIOS POR MATERIA
    ws_materias = wb.create_sheet("Promedios por Materia")
    ws_materias['A1'] = 'Materia'
    ws_materias['B1'] = 'Promedio'

    for cell in ['A1', 'B1']:
        ws_materias[cell].fill = header_fill
        ws_materias[cell].font = header_font

    for idx, (materia, promedio) in enumerate(datos_reporte['promedios_materias'], start=2):
        ws_materias[f'A{idx}'] = materia
        ws_materias[f'B{idx}'] = round(promedio, 2)

    # Agregar gráfico de barras
    if datos_reporte['promedios_materias']:
        chart = BarChart()
        chart.title = "Promedios por Materia"
        chart.x_axis.title = "Materia"
        chart.y_axis.title = "Promedio"

        data = Reference(ws_materias, min_col=2, min_row=1,
                        max_row=len(datos_reporte['promedios_materias']) + 1)
        categories = Reference(ws_materias, min_col=1, min_row=2,
                              max_row=len(datos_reporte['promedios_materias']) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws_materias.add_chart(chart, "D2")

    # HOJA 3: TOP ALUMNOS
    ws_alumnos = wb.create_sheet("Top Alumnos")

    # Sección promedios
    ws_alumnos['A1'] = 'TOP ALUMNOS POR PROMEDIO'
    ws_alumnos['A1'].font = Font(bold=True, size=12)
    ws_alumnos['A2'] = 'Alumno'
    ws_alumnos['B2'] = 'Promedio'

    for cell in ['A2', 'B2']:
        ws_alumnos[cell].fill = header_fill
        ws_alumnos[cell].font = header_font

    row = 3
    for alumno, promedio in datos_reporte['alumnos_top_promedio']:
        ws_alumnos[f'A{row}'] = alumno
        ws_alumnos[f'B{row}'] = round(promedio, 2)
        row += 1

    # Sección asistencias
    row += 2
    ws_alumnos[f'A{row}'] = 'TOP ALUMNOS POR ASISTENCIA'
    ws_alumnos[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    ws_alumnos[f'A{row}'] = 'Alumno'
    ws_alumnos[f'B{row}'] = 'Asistencia (%)'

    for cell in [f'A{row}', f'B{row}']:
        ws_alumnos[cell].fill = header_fill
        ws_alumnos[cell].font = header_font

    row += 1
    for alumno, asistencia in datos_reporte['alumnos_top_asistencia']:
        ws_alumnos[f'A{row}'] = alumno
        ws_alumnos[f'B{row}'] = round(asistencia, 2)
        row += 1

    # Ajustar anchos de columna
    for ws in [ws_resumen, ws_materias, ws_alumnos]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
